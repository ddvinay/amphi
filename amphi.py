import openpyxl
import urllib2
import time
import sys
import datetime

amfiHyperlink = 'http://portal.amfiindia.com/DownloadNAVHistoryReport_Po.aspx?mf=%d&tp=1&frmdt=%s&todt=%s'

fundHouseToCodes = dict()
fundHouseToCodes['Birla Sun Life Mutual Fund'] = 3
fundHouseToCodes['Baroda Pioneer Mutual Fund'] = 4
fundHouseToCodes['DSP BlackRock Mutual Fund']  = 6
fundHouseToCodes['HDFC Mutual Fund'] = 9
fundHouseToCodes['PRINCIPAL Mutual Fund'] = 10
fundHouseToCodes['Escorts Mutual Fund'] = 13
fundHouseToCodes['JM Financial Mutual Fund'] = 16
fundHouseToCodes['Kotak Mahindra Mutual Fund'] = 17
fundHouseToCodes['LIC Mutual Fund'] = 18
fundHouseToCodes['ICICI Prudential Mutual Fund'] = 20
fundHouseToCodes['Reliance Mutual Fund'] = 21
fundHouseToCodes['SBI Mutual Fund'] = 22
fundHouseToCodes['Tata Mutual Fund'] = 25
fundHouseToCodes['Taurus Mutual Fund'] = 26
fundHouseToCodes['Franklin Templeton Mutual Fund'] = 27
fundHouseToCodes['UTI Mutual Fund'] = 28
fundHouseToCodes['Canara Robeco Mutual Fund'] = 32
fundHouseToCodes['Sundaram Mutual Fund'] = 33
fundHouseToCodes['Sahara Mutual Fund'] = 35
fundHouseToCodes['HSBC Mutual Fund'] = 37
fundHouseToCodes['Quantum Mutual Fund'] = 41
fundHouseToCodes['Invesco Mutual Fund'] = 42
fundHouseToCodes['Mirae Asset Mutual Fund'] = 45
fundHouseToCodes['BOI AXA Mutual Fund'] = 46
fundHouseToCodes['Edelweiss Mutual Fund'] = 47
fundHouseToCodes['IDFC Mutual Fund'] = 48
fundHouseToCodes['Axis Mutual Fund'] = 53
fundHouseToCodes['Peerless Mutual Fund'] = 54
fundHouseToCodes['Motilal Oswal Mutual Fund'] = 55
fundHouseToCodes['L&T Mutual Fund'] = 56
fundHouseToCodes['IDBI Mutual Fund'] = 57
fundHouseToCodes['DHFL Pramerica Mutual Fund'] = 58
fundHouseToCodes['BNP Paribas Mutual Fund'] = 59
fundHouseToCodes['Union Mutual Fund'] = 61
fundHouseToCodes['IIFL Mutual Fund'] = 62
fundHouseToCodes['Indiabulls Mutual Fund'] = 63
fundHouseToCodes['PPFAS Mutual Fund'] = 64
fundHouseToCodes['Shriram Mutual Fund'] = 67
fundHouseToCodes['Mahindra Mutual Fund'] = 69

def printFundHouseCodes():
    for i in range(100):
        response = urllib2.urlopen(amfiHyperlink % (i, '1-Mar-2017', '1-Mar-2017'))
        rawData = response.read().split('\n')
        fltrdData = rawData[0].replace('\n', '').replace(' ', '').replace('\b', '').replace('\t', '').replace('\r', '')
        if fltrdData:
            print '\'%s\' = %d,' % (rawData[5].strip(), i)
        sys.stdout.flush()

def getHistoricNAV(fundHouse, schemeCode, transactionDate):

    fundHouseCode = fundHouseToCodes[fundHouse]
    response = urllib2.urlopen(amfiHyperlink % (fundHouseCode, transactionDate, transactionDate))
    
    rawData = response.read().split('\n')

    historicNAV = ''
    for line in rawData:
        if line.find(schemeCode) != -1:
            historicNAV = line.split(';')[2]

    if historicNAV != '':
        return historicNAV
    else:
        print 'Failed to fetch the historic value'
        print 'Link used:',
        print amfiHyperlink % (fundHouseCode, transactionDate, transactionDate)

def updateTransactionsSheet(passbook):

    portfolioSheet = passbook.get_sheet_by_name('portfolio')
    transactionsSheet = passbook.get_sheet_by_name('transactions')

    print 'Loading portfolio ...',
    nameToDetails = dict()
    i = 0
    while True:
        schemeName = portfolioSheet['D%d' % (i+2)].value
        if schemeName == None:
            break
        fundHouse  = portfolioSheet['C%d' % (i+2)].value
        schemeCode = str(portfolioSheet['A%d' % (i+2)].value)
        nameToDetails[schemeName] = (fundHouse, schemeCode)
        i += 1
    print 'Done.'

    sys.stdout.flush()

    i = 0
    while True:
        schemeName = transactionsSheet['A%d' % (i+2)].value
        if schemeName == None:
            break
        if (transactionsSheet['E%d' % (i+2)].value == 0.0 or
            transactionsSheet['E%d' % (i+2)].value == None):
            print 'Updating transaction # %d' % (i+2)
            (fundHouse, schemeCode)  = nameToDetails[schemeName]
            transactionDateRaw = transactionsSheet['B%d' % (i+2)].value
            transactionDate = transactionDateRaw.date().strftime('%d-%b-%y')      
            print schemeName, transactionDate, fundHouse, schemeCode
            NAV = getHistoricNAV(fundHouse, schemeCode, transactionDate)
            transactionsSheet['E%d' % (i+2)].value = float(NAV)
        i += 1

def updateNAV(passbook):
    portfolioSheet = passbook.get_sheet_by_name('portfolio')

    isinCodes = list()

    i = 0
    while True:
        isinCode = portfolioSheet['B%d' % (i+2)].value
        if isinCode == None:
            break;
        isinCodes.append(isinCode)
        i = i + 1

    print 'Donwloading latest NAVs ... ',
    sys.stdout.flush()

    response = urllib2.urlopen('http://portal.amfiindia.com/spages/NAV0.txt')
    rawData = response.read().split('\n')

    print 'Done'
    sys.stdout.flush()

    NAV = dict()
    NAVDate = dict()

    # Extract NAVs
    for line in rawData:
        for isinCode in isinCodes:
            if line.find(isinCode) != -1:
                NAV[isinCode] = line.split(';')[4].strip()
                dateRaw = line.split(';')[7].strip()
                NAVDate[isinCode] = datetime.datetime.strptime(dateRaw, '%d-%b-%Y').date()

    # Update spreadsheet
    i = 0

    while True:
        isinCode = portfolioSheet['B%d' % (i+2)].value
        if isinCode == None:
            break
        print '%s # %s (%s)' % (isinCode, NAV[isinCode], NAVDate[isinCode])

        portfolioSheet['E%d' % (i+2)] = float(NAV[isinCode])
        portfolioSheet['F%d' % (i+2)] = NAVDate[isinCode]
        portfolioSheet['F%d' % (i+2)].number_format = 'dd-mmm-yy'
        i = i + 1

def main():
    command    = sys.argv[1]
    inputFile  = sys.argv[2]
    outputFile = sys.argv[3]

    
    
    if command == 'update-tr':
        passbook = openpyxl.load_workbook(inputFile)
        updateTransactionsSheet(passbook)
        passbook.save(outputFile)

    if command == 'update-nav':
        passbook = openpyxl.load_workbook(inputFile)
        updateNAV(passbook)
        passbook.save(outputFile)

    if command == 'update-both':
        passbook = openpyxl.load_workbook(inputFile)
        updateTransactionsSheet(passbook)
        updateNAV(passbook)
        passbook.save(outputFile)

    if command == 'print-codes':
        printFundHouseCodes()        
     
main()
